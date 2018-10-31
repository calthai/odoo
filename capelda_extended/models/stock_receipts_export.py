# -*- coding: utf-8 -*-
# Copyright (C) 2016-2017  Technaureus Info Solutions(<http://technaureus.com/>).
import xlsxwriter
from openpyxl import writer

from odoo import models, fields, api, _
from datetime import datetime
from StringIO import StringIO
import xlwt
import time
import xlsxwriter
import base64
from odoo.exceptions import UserError
from odoo.tools import misc


# this is for tax report section
class stock_receipts_report(models.TransientModel):
    _name = 'stock.receipts.report'

    # type_report = fields.Char(string='ประเภทรายงาน:')
    # date_from = fields.Date(string='ตั้งแต่วันที่:')
    # date_to = fields.Date(string='ถึงวันที่:')
    # picking_no = fields.Many2one('stock.picking', string='Name')
    stock_receipt_ids = fields.Many2many('stock.picking', string='Receipt Name')


    # @api.model
    # def default_get(self, fields):
    #     localtime = time.localtime(time.time())
    #     # mm = localtime.tm_min tm_year tm_mon	1 to 12 tm_mday
    #     res = super(stock_receipts_report, self).default_get(fields)
    #     curr_date = datetime.now()
    #     from_date = datetime(curr_date.year, curr_date.month, 01).date() or False
    #     to_date = datetime(curr_date.year, curr_date.month, curr_date.day ).date() or False
    #     res.update({'date_from': str(from_date), 'date_to': str(to_date)})
    #     # from_date = datetime(localtime.tm_year, localtime.tm_mon, 01).date() or False
    #     # to_date = datetime(localtime.tm_year, localtime.tm_mon, localtime.tm_mday).date() or False
    #     # res.update({'date_from': str(from_date), 'date_to': str(to_date)})
    #     return res


    @api.multi
    def print_file(self):
        fl = StringIO()
        workbook = xlwt.Workbook(encoding='utf-8')
        # workbook = xlsxwriter.Workbook('filename.xlsx')
        # workbooks = xlsxwriter.Workbook('filename.xlsx')
        # workbook = xlwt.Workbook('filename.xlsx')
        font = xlwt.Font()
        font.bold = True
        font.bold = True
        for_right = xlwt.easyxf(
            "font: name  Times New Roman,color black,  height 180;  align: horiz right,vertical center; borders: top thin, bottom thin, left thin, right thin")
        for_right.num_format_str = '#,###.00'
        for_right_bold = xlwt.easyxf(
            "font: bold 1, name  Times New Roman,color black,  height 180;  align: horiz right,vertical center; borders: top thin, bottom thin, left thin, right thin")
        for_right_bold.num_format_str = '#,###.00'
        for_center = xlwt.easyxf(
            "font: name  Times New Roman, color black,  height 180; align: horiz center,vertical center,wrap on; borders: top thin, bottom thin, left thin, right thin")
        for_left = xlwt.easyxf(
            "font: name  Times New Roman,color black,  height 180;  align: horiz left,vertical center; borders: top thin, bottom thin, left thin, right thin")
        for_center_bold = xlwt.easyxf(
            "font: bold 1, name  Times New Roman, color black, height 180;  align: horiz center,vertical center,wrap on; borders: top thin, bottom thin, left thin, right thin")
        for_left_bold_no_border = xlwt.easyxf(
            "font: bold 1, name  Times New Roman, color black, height 180;  align: horiz left,vertical center;")

        GREEN_TABLE_HEADER = xlwt.easyxf(
            'font: bold 1, name  Times New Roman, height 300,color black;'
            'align: vertical center, horizontal center, wrap on;'
            'borders: top thin, bottom thin, left thin, right thin;'
            'pattern:  pattern_fore_colour white, pattern_back_colour white'
        )

        alignment = xlwt.Alignment()  # Create Alignment
        alignment.horz = xlwt.Alignment.HORZ_RIGHT
        style = xlwt.easyxf('align: wrap yes')
        style.num_format_str = '#,###.00'
        cr, uid, context = self.env.args
        #

        i = 0
        domain = [('state', '=', 'done'), ('picking_type_code', '=', 'incoming')]
        if self.stock_receipt_ids:
            domain.append(('id', 'in', self.stock_receipt_ids.ids))
        stock_receipt_id = self.env['stock.picking'].search(domain)
        # stock_receipt_id = self.env['stock.picking'].search([('state', '=', 'done'), ('picking_type_code', '=', 'incoming')])

        if stock_receipt_id:

            worksheet = workbook.add_sheet('CAPELDA')
            worksheet.row(0).height = 200
            worksheet.col(0).width = 4000
            worksheet.col(1).width = 4000
            worksheet.col(2).width = 4000
            worksheet.col(3).width = 4000
            worksheet.col(4).width = 4000
            worksheet.col(5).width = 20000
            worksheet.col(6).width = 4000
            worksheet.col(7).width = 4000
            worksheet.col(8).width = 4000
            worksheet.col(9).width = 4000
            worksheet.col(10).width = 1500
            worksheet.col(11).width = 1500
            worksheet.col(12).width = 1500
            worksheet.col(13).width = 1500
            worksheet.col(14).width = 1500
            worksheet.col(15).width = 3000
            worksheet.col(16).width = 3000
            worksheet.col(17).width = 3000
            worksheet.col(18).width = 3000
            worksheet.col(19).width = 3000
            worksheet.col(20).width = 3000
            worksheet.col(21).width = 3000
            worksheet.col(22).width = 3000
            worksheet.col(23).width = 3000
            worksheet.col(24).width = 3000
            worksheet.col(25).width = 3000
            worksheet.col(26).width = 3000
            worksheet.col(27).width = 3000
            worksheet.col(28).width = 3000
            worksheet.col(29).width = 3000
            worksheet.col(30).width = 3000
            worksheet.col(31).width = 3000
            worksheet.col(32).width = 3000

            borders = xlwt.Borders()
            borders.bottom = xlwt.Borders.MEDIUM
            border_style = xlwt.XFStyle()  # Create Style
            border_style.borders = borders

            inv_row = 3

            worksheet.write_merge(0, 0, 0, 7, 'Calpeda (Thailand) Co. Ltd.', for_center_bold)
            worksheet.write_merge(1, 1, 0, 7, 'Stock Receipt Export', for_center_bold)
            # worksheet.write_merge(1, 1, 0, 32, 'Product Export', for_center_bold)
            # worksheet.write(0, 3,  'AGT CO.,LTD.', for_center_bold)
            # worksheet.write(1, 3,  'Product Export', for_center_bold)

            # The Mall Code , QTY ,Sku ,Des ,GP ,MAP ,Ex ,vat

            worksheet.write(2, 0, 'NAME', for_center_bold)
            worksheet.write(2, 1, 'SOURCE DOCUMENT', for_center_bold)
            worksheet.write(2, 2, 'DESTINATION LOCATION', for_center_bold)
            worksheet.write(2, 3, 'NEW SOURCE LOCATION', for_center_bold)
            worksheet.write(2, 4, 'DESTINATION LOCATION', for_center_bold)
            worksheet.write(2, 5, 'PRODUCT', for_center_bold)
            worksheet.write(2, 6, 'DONE', for_center_bold)

            for stock_receipt_ids in stock_receipt_id:

                for stock_receipt_line_ids in stock_receipt_ids.move_lines:

                    i += 1

                    worksheet.write(inv_row, 0, str(stock_receipt_ids.name), for_center)
                    worksheet.write(inv_row, 1, str(stock_receipt_ids.origin), for_center)
                    worksheet.write(inv_row, 2, str(stock_receipt_ids.location_dest_id.name), for_center)
                    if stock_receipt_line_ids.new_location_id:
                        worksheet.write(inv_row, 3, str(stock_receipt_line_ids.new_location_id.name), for_center)
                    else:
                        worksheet.write(inv_row, 3, '', for_center)
                    worksheet.write(inv_row, 4, str(stock_receipt_line_ids.location_dest_id.name), for_center)
                    worksheet.write(inv_row, 5, str(stock_receipt_line_ids.product_id.name), for_center)

                    worksheet.write(inv_row, 6, stock_receipt_line_ids.quantity_done, for_center)

                    inv_row += 1

                    # ------------------------#

        elif not stock_receipt_id:

            raise UserError(_('There is record this date range.'))

        workbook.save(fl)

        fl.seek(0)

        buf = base64.encodestring(fl.read())
        cr, uid, context = self.env.args
        ctx = dict(context)
        ctx.update({'report_file': buf})
        self.env.args = cr, uid, misc.frozendict(context)

        ## To remove those previous saved report data from table. To avoid unwanted storage
        self._cr.execute("TRUNCATE stock_picking_export CASCADE")

        wizard_id = self.env['stock.picking.export'].create(
            vals={'name': 'Stock Receiving Export.xls', 'report_file': ctx['report_file']})

        return {
            'type': 'ir.actions.act_window',
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'stock.picking.export',
            'target': 'new',
            'context': ctx,
            'res_id': wizard_id.id,
        }


        # workbook.save(fl)
        #         fl.seek(0)
        #
        #         buf = base64.encodestring(fl.read())
        #         cr, uid, context = self.env.args
        #         ctx = dict(context)
        #         ctx.update({'report_file': buf})
        #         self.env.args = cr, uid, misc.frozendict(context)
        #         ## To remove those previous saved report data from table. To avoid unwanted storage
        #         self._cr.execute("TRUNCATE stock_inventory_export CASCADE")
        #         wizard_id = self.env['stock.inventory.export'].create(
        #             vals={'name': 'Stock Report.xls', 'report_file': ctx['report_file']})
        #         # print wizard_id
        #         return {
        #             'type': 'ir.actions.act_window',
        #             'view_type': 'form',
        #             'view_mode': 'form',
        #             'res_model': 'stock.inventory.export',
        #             'target': 'new',
        #             'context': ctx,
        #             'res_id': wizard_id.id,
        #         }


class stock_picking_export(models.TransientModel):
    _name = 'stock.picking.export'

    report_file = fields.Binary('File')
    name = fields.Char(string='File Name', size=32)

    @api.multi
    def action_back_export(self):
        if self._context is None:
            self._context = {}
        return {
            'type': 'ir.actions.act_window',
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'stock.receipts.report',
            'target': 'new',
        }

